# 1、遍历所有月份，构建一个{姓名:{月份:领取金额}}的字典
# 如{'张三': {'1月份':100,......,'总计': 800000}}
# 2、将字典写到excel中
from openpyxl import Workbook
from openpyxl import load_workbook
# import
amount_dict = {}
month_list = ['1月', '2月', '3月']
# month_list = ['1月']
for month in month_list:
    # 遍历打开各个月份的excel
    excel_file = '生活(%s).xlsx' %(month)
    wb = load_workbook(excel_file) # 逐个打开excel
    # 遍历进去每个页签
    # print(wb.sheetnames)  ['城镇', '农村', '集中']
    for sheet_name in wb.sheetnames: # 逐个打开sheet
        # sheet = wb.get_sheet_by_name(sheet_name)
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row[3:6:2]:
                cell_value = cell.value
                if cell_value == None:
                    break
                if str(cell_value) in ['总额', '金 额' ,'户 名', '金额', None, '复核：'] or '=SUM' in str(cell_value):
                    break
                if cell.column == 4:
                    name = cell.value
                    # amount_dict[name] = {}
                if cell.column == 6:
                    amount = cell.value
                    if name not in amount_dict:
                        amount_dict[name] = {}
                        amount_dict[name][month] = amount
                    else:
                        amount_dict[name][month] = amount
                    # amount_dict[name] = amount

def total(dict1):
    sum = 0
    for k in dict1:
        sum += dict1[k]
    return sum
for name2 in amount_dict:
    amount2 = amount_dict[name2]
    heji = total(amount2)
    amount_dict[name2]['合计'] = heji

# 优化下amount_dict,某月份没发的置为0

# 插入数据

list_head = ['姓名','1月', '2月', '3月','4月', '5月', '6月', '7月', '8月', '9月','10月', '11月', '12月', '合计']
wb_write = Workbook()
ws_write = wb_write.active
# ws_write.append(list_head)
for x in amount_dict.items():
    list1 = []
    name3 = x[0]
    amount3 = x[1] # 子字典
    if len(amount3) == 4:
        continue
    list1.append(name3)


    for yue in amount3:
        jine3 = amount3[yue]
        ab = yue + ':' + str(jine3)
        list1.append(ab)
    ws_write.append(list1)
wb_write.save('result111111.xlsx')


    # for m in month_list:
    #     if m in amount3:
    #         continue
    #     else:
    #         amount3[m] = 0
    #         sorted(amount3.keys())
